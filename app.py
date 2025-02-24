import streamlit as st
import pandas as pd
from database import DatabaseManager
from datetime import datetime, timedelta
import calendar
import plotly.express as px
import plotly.graph_objects as go
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import numpy as np
import os

class ISPStreamlitApp:
    def __init__(self):
        self.db = DatabaseManager()
        st.set_page_config(page_title="ISP Management System", layout="wide")
        
    def run(self):
        st.title("ISP Management System")
        
        # Sidebar navigation
        page = st.sidebar.selectbox(
            "Select Page",
            ["Monthly Payments", "Customer Management", "Import Data", "Analytics", "Reports"]
        )
        
        if page == "Monthly Payments":
            self.show_monthly_payments_page()
        elif page == "Customer Management":
            self.show_customer_management()
        elif page == "Import Data":
            self.show_import_page()
        elif page == "Analytics":
            self.show_analytics_page()
        elif page == "Reports":
            self.show_reports_page()

    # ... [previous methods remain unchanged until show_import_page] ...

    def show_import_page(self):
        st.header("Import Data")
        
        uploaded_file = st.file_uploader("Choose Excel file", type=['xlsx', 'xls'])
        
        if uploaded_file is not None:
            try:
                # Preview data
                df = pd.read_excel(uploaded_file)
                st.subheader("Data Preview")
                st.dataframe(df.head(10))
                
                # Show statistics
                st.subheader("Data Statistics")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Records", len(df))
                with col2:
                    st.metric("Unique Names", df['name'].nunique())
                with col3:
                    st.metric("Total Value", f"${df['monthly_value'].sum():,.2f}")
                
                # Import button
                if st.button("Import Data"):
                    # Save uploaded file temporarily
                    temp_path = os.path.join('data', 'temp_import.xlsx')
                    with open(temp_path, 'wb') as f:
                        f.write(uploaded_file.getvalue())
                    
                    # Import data
                    success, message = self.db.import_excel_data(temp_path)
                    
                    # Remove temporary file
                    os.remove(temp_path)
                    
                    if success:
                        st.success(message)
                    else:
                        st.error(message)
                
            except Exception as e:
                st.error(f"Error reading file: {str(e)}")
                st.error("Please make sure your Excel file has the correct columns")
                st.info("""
                Required columns:
                - name
                - address
                - phone
                - mbps
                - state
                - contract_date
                - payment_day
                - payment_type
                - bank
                - iban
                - monthly_value
                - expiration_date
                """)

    def process_expiration_data(self, file):
        df = pd.read_excel(file)
        
        # Validate required columns
        required_columns = ['customer_id', 'expiration_date', 'service_type']
        if not all(col in df.columns for col in required_columns):
            st.error("Excel file must contain: customer_id, expiration_date, service_type")
            return
        
        # Convert dates to datetime
        df['expiration_date'] = pd.to_datetime(df['expiration_date'])
        current_date = datetime.now()
        
        # Create expiration categories
        df['days_to_expire'] = (df['expiration_date'] - current_date).dt.days
        df['status'] = pd.cut(
            df['days_to_expire'],
            bins=[-np.inf, 0, 30, 90, np.inf],
            labels=['Expired', 'Critical', 'Warning', 'OK']
        )
        
        # Display summary
        st.subheader("Expiration Summary")
        summary = df['status'].value_counts()
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Expired", summary.get('Expired', 0))
        with col2:
            st.metric("Critical (< 30 days)", summary.get('Critical', 0))
        with col3:
            st.metric("Warning (< 90 days)", summary.get('Warning', 0))
        with col4:
            st.metric("OK", summary.get('OK', 0))
        
        # Display detailed data
        st.dataframe(df)
        
        # Generate report
        if st.button("Generate Expiration Report"):
            report_buffer = self.generate_expiration_report(df)
            st.download_button(
                label="Download Expiration Report",
                data=report_buffer,
                file_name="expiration_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    def process_payment_data(self, file):
        df = pd.read_excel(file)
        
        # Validate required columns
        required_columns = ['customer_id', 'payment_date', 'amount', 'status']
        if not all(col in df.columns for col in required_columns):
            st.error("Excel file must contain: customer_id, payment_date, amount, status")
            return
        
        # Process payment data
        df['payment_date'] = pd.to_datetime(df['payment_date'])
        df['month'] = df['payment_date'].dt.month
        df['year'] = df['payment_date'].dt.year
        
        # Display summary by month
        st.subheader("Payment Summary by Month")
        monthly_summary = df.groupby(['year', 'month']).agg({
            'amount': 'sum',
            'customer_id': 'count'
        }).reset_index()
        
        st.dataframe(monthly_summary)
        
        # Generate payment report
        if st.button("Generate Payment Report"):
            report_buffer = self.generate_payment_report(df)
            st.download_button(
                label="Download Payment Report",
                data=report_buffer,
                file_name="payment_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    def generate_expiration_report(self, df):
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write summary sheet
            summary = df['status'].value_counts().reset_index()
            summary.columns = ['Status', 'Count']
            summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # Write detailed sheet
            df_sorted = df.sort_values('days_to_expire')
            df_sorted.to_excel(writer, sheet_name='Detailed', index=False)
            
            # Format workbook
            workbook = writer.book
            
            # Format summary sheet
            summary_sheet = workbook['Summary']
            for row in summary_sheet.iter_rows(min_row=2):
                status = row[0].value
                if status == 'Expired':
                    for cell in row:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                elif status == 'Critical':
                    for cell in row:
                        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            
        output.seek(0)
        return output.getvalue()

    def generate_payment_report(self, df):
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write monthly summary
            monthly_summary = df.groupby(['year', 'month']).agg({
                'amount': ['sum', 'count'],
                'status': lambda x: (x == 'Paid').sum()
            }).reset_index()
            monthly_summary.columns = ['Year', 'Month', 'Total Amount', 'Total Payments', 'Paid Count']
            monthly_summary.to_excel(writer, sheet_name='Monthly Summary', index=False)
            
            # Write detailed data
            df.to_excel(writer, sheet_name='Detailed', index=False)
        
        output.seek(0)
        return output.getvalue()

    def show_reports_page(self):
        st.header("Reports")
        
        report_type = st.selectbox(
            "Select Report Type",
            ["Expiration Report", "Payment Report", "Customer Status Report"]
        )
        
        col1, col2 = st.columns(2)
        with col1:
            start_date = st.date_input("Start Date", value=datetime.now() - timedelta(days=30))
        with col2:
            end_date = st.date_input("End Date", value=datetime.now())
        
        if st.button("Generate Report"):
            if report_type == "Expiration Report":
                expirations = self.db.get_expirations(start_date, end_date)
                if expirations:
                    df = pd.DataFrame(expirations)
                    self.show_expiration_analysis(df)
            elif report_type == "Payment Report":
                payments = self.db.get_payments(start_date, end_date)
                if payments:
                    df = pd.DataFrame(payments)
                    self.show_payment_analysis(df)
            else:
                customers = self.db.get_customer_status()
                if customers:
                    df = pd.DataFrame(customers)
                    self.show_customer_analysis(df)

    def show_expiration_analysis(self, df):
        st.subheader("Expiration Analysis")
        
        # Create expiration categories
        df['days_to_expire'] = (pd.to_datetime(df['expiration_date']) - datetime.now()).dt.days
        
        # Show statistics
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Expirations", len(df))
        with col2:
            expired = len(df[df['days_to_expire'] < 0])
            st.metric("Expired", expired)
        with col3:
            critical = len(df[(df['days_to_expire'] >= 0) & (df['days_to_expire'] <= 30)])
            st.metric("Critical (Next 30 days)", critical)
        
        # Show expiration timeline
        fig = px.timeline(
            df,
            x_start='contract_date',
            x_end='expiration_date',
            y='customer_id',
            title="Contract Timeline"
        )
        st.plotly_chart(fig)

    def show_payment_analysis(self, df):
        st.subheader("Payment Analysis")
        
        # Show payment statistics
        total_amount = df['amount'].sum()
        paid_amount = df[df['status'] == 'Paid']['amount'].sum()
        payment_rate = (paid_amount / total_amount) * 100 if total_amount > 0 else 0
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Amount", f"${total_amount:,.2f}")
        with col2:
            st.metric("Paid Amount", f"${paid_amount:,.2f}")
        with col3:
            st.metric("Payment Rate", f"{payment_rate:.1f}%")
        
        # Show payment trend
        fig = px.line(
            df.groupby('payment_date')['amount'].sum().reset_index(),
            x='payment_date',
            y='amount',
            title="Daily Payment Trend"
        )
        st.plotly_chart(fig)

    def show_customer_analysis(self, df):
        st.subheader("Customer Status Analysis")
        
        # Show customer statistics
        total_customers = len(df)
        active_customers = len(df[df['status'] == 'Active'])
        inactive_rate = ((total_customers - active_customers) / total_customers) * 100 if total_customers > 0 else 0
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Customers", total_customers)
        with col2:
            st.metric("Active Customers", active_customers)
        with col3:
            st.metric("Inactive Rate", f"{inactive_rate:.1f}%")
        
        # Show customer distribution
        fig = px.pie(
            df,
            names='status',
            title="Customer Status Distribution"
        )
        st.plotly_chart(fig)

    def show_customer_management(self):
        st.header("Customer Management")
        
        # Tabs for different sections
        tab1, tab2, tab3 = st.tabs(["Add New Customer", "View All Customers", "Manage Customer Status"])
        
        # Add New Customer Tab
        with tab1:
            with st.form("new_customer"):
                col1, col2 = st.columns(2)
                with col1:
                    name = st.text_input("Customer Name")
                    phone = st.text_input("Phone")
                    mbps = st.number_input("Mbps", min_value=1)
                    state = st.selectbox(
                        "Status",
                        options=["Active", "Inactive"],
                        index=0
                    )
                    contract_date = st.date_input(
                        "Contract Start Date",
                        value=datetime.now(),
                        min_value=datetime(2020, 1, 1),
                        max_value=datetime(2030, 12, 31)
                    )
                
                with col2:
                    payment_type = st.selectbox(
                        "Payment Type",
                        ["Bank Transfer", "Credit Card", "Direct Debit"]
                    )
                    bank = st.text_input("Bank")
                    iban = st.text_input("IBAN")
                    value = st.number_input("Monthly Value", min_value=0.0)
                    address = st.text_input("Address")
                    payment_day = st.number_input("Payment Day of Month", min_value=1, max_value=31, value=1)
                
                if st.form_submit_button("Add Customer"):
                    try:
                        customer_data = {
                            'name': name,
                            'address': address,
                            'phone': phone,
                            'mbps': mbps,
                            'state': state,
                            'contract_date': contract_date.strftime('%Y-%m-%d'),
                            'payment_day': payment_day
                        }
                        
                        payment_data = {
                            'payment_type': payment_type,
                            'bank': bank,
                            'iban': iban,
                            'value': value,
                            'expiration_date': (contract_date + timedelta(days=365)).strftime('%Y-%m-%d')
                        }
                        
                        success, result = self.db.register_customer(customer_data, payment_data)
                        if success:
                            st.success("Customer added successfully!")
                        else:
                            st.error(f"Failed to add customer: {result}")
                    except Exception as e:
                        st.error(f"Error: {str(e)}")
        
        # View All Customers Tab
        with tab2:
            try:
                customers = self.db.get_all_customers()
                if customers:
                    df = pd.DataFrame(customers, columns=[
                        'ID', 'Name', 'Address', 'Phone', 'Mbps', 
                        'Status', 'Contract Date', 'Payment Day', 'Payment Type', 
                        'Bank', 'Monthly Value', 'Last Payment Date', 'Payment Status'
                    ])
                    
                    # Filters
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        status_filter = st.multiselect(
                            "Filter by Status",
                            options=df['Status'].unique(),
                            default=df['Status'].unique()
                        )
                    with col2:
                        payment_status_filter = st.multiselect(
                            "Filter by Payment Status",
                            options=df['Payment Status'].unique(),
                            default=df['Payment Status'].unique()
                        )
                    with col3:
                        search_name = st.text_input("Search by Name")
                    
                    # Apply filters
                    mask = (
                        df['Status'].isin(status_filter) &
                        df['Payment Status'].isin(payment_status_filter)
                    )
                    if search_name:
                        mask = mask & df['Name'].str.contains(search_name, case=False)
                    
                    filtered_df = df[mask]
                    
                    # Statistics
                    st.subheader("Customer Statistics")
                    stats_col1, stats_col2, stats_col3, stats_col4 = st.columns(4)
                    with stats_col1:
                        st.metric("Total Customers", len(filtered_df))
                    with stats_col2:
                        active_count = len(filtered_df[filtered_df['Status'] == 'Active'])
                        st.metric("Active Customers", active_count)
                    with stats_col3:
                        pending_payments = len(filtered_df[filtered_df['Payment Status'] == 'Pending'])
                        st.metric("Pending Payments", pending_payments)
                    with stats_col4:
                        overdue_payments = len(filtered_df[filtered_df['Payment Status'] == 'Overdue'])
                        st.metric("Overdue Payments", overdue_payments)
                    
                    # Customer Table
                    st.subheader("Customer List")
                    st.dataframe(
                        filtered_df,
                        hide_index=True,
                        use_container_width=True
                    )
                    
                    # Export option
                    csv = filtered_df.to_csv(index=False).encode('utf-8')
                    st.download_button(
                        "Export Customer List",
                        csv,
                        "customers.csv",
                        "text/csv",
                        key='export-customers'
                    )
                else:
                    st.info("No customers found in the database.")
            except Exception as e:
                st.error(f"Error loading customers: {str(e)}")
        
        # Manage Customer Status Tab
        with tab3:
            try:
                customers = self.db.get_all_customers()
                if customers:
                    df = pd.DataFrame(customers, columns=[
                        'ID', 'Name', 'Status', 'Payment Status', 'Last Payment Date'
                    ])
                    
                    st.subheader("Manage Customer Status")
                    
                    # Customer selection
                    customer_id = st.selectbox(
                        "Select Customer",
                        options=df['ID'].tolist(),
                        format_func=lambda x: f"{df[df['ID'] == x]['Name'].iloc[0]} (ID: {x})"
                    )
                    
                    if customer_id:
                        customer_data = df[df['ID'] == customer_id].iloc[0]
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            st.info(f"Current Status: {customer_data['Status']}")
                            if st.button(
                                "Activate" if customer_data['Status'] == 'Inactive' else "Deactivate",
                                key=f"toggle_{customer_id}"
                            ):
                                new_status = 'Active' if customer_data['Status'] == 'Inactive' else 'Inactive'
                                success = self.db.update_customer_status(customer_id, new_status)
                                if success:
                                    st.success(f"Customer status updated to {new_status}")
                                    st.rerun()
                                else:
                                    st.error("Failed to update customer status")
                        
                        with col2:
                            st.info(f"Payment Status: {customer_data['Payment Status']}")
                            if customer_data['Payment Status'] == 'Overdue':
                                if st.button("Mark as Paid", key=f"pay_{customer_id}"):
                                    success = self.db.record_payment(customer_id)
                                    if success:
                                        st.success("Payment recorded successfully")
                                        st.rerun()
                                    else:
                                        st.error("Failed to record payment")
                
                else:
                    st.info("No customers found in the database.")
            except Exception as e:
                st.error(f"Error managing customer status: {str(e)}")

if __name__ == "__main__":
    app = ISPStreamlitApp()
    app.run()